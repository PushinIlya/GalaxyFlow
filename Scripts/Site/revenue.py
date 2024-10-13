#!/usr/bin/env python
# coding: utf-8

# ## Формирование отчёта о финансовых результатах

# ### Расчёт "Выручка"

# In[ ]:


'''
Скрипт для преобразования исходных данных из 1C
'''
# импортируем библиотеки
import pandas as pd
import numpy as np
import os
import pathlib
from openpyxl import load_workbook
from datetime import datetime, timedelta


# In[ ]:


# запишем путь к папке с исходными данными в переменную
source = pathlib.Path(os.getcwd().replace('Fin_Flow', os.path.join('YandexDisk', 'Fin_Flow')).replace('Scripts', 'Source data'))


# In[ ]:


# создадим списки с названиями всех необходимых файлов в папке с исходными данными
source_rev_name = [raw_data for raw_data in os.listdir(source) if 'Выручка' in raw_data and
                   raw_data.endswith('.xlsx')]
source_rev_exceptions_name = [raw_data for raw_data in os.listdir(source) if 'НДС' in raw_data and
                              raw_data.endswith('.xlsx')]

print(source_rev_name)
print(source_rev_exceptions_name)


# In[ ]:


# загрузим excel-файлы в pandas
rev = {raw_data: pd.read_excel(os.path.join(source, f'{raw_data}'), header=None) for raw_data in source_rev_name}
rev_exceptions = {raw_data: pd.read_excel(os.path.join(source, f'{raw_data}'), header=None) for raw_data in source_rev_exceptions_name}


# In[ ]:


# загрузим excel-файлы в openpyxl
rev_structure = {raw_data: load_workbook(filename=os.path.join(source, f'{raw_data}')) for raw_data in source_rev_name}
rev_exceptions_structure = {raw_data: load_workbook(filename=os.path.join(source, f'{raw_data}')) for raw_data in source_rev_exceptions_name}


# In[ ]:


# получим первый лист в каждом excel-файле
rev_structure = {i[0]: i[1].worksheets[0] for i in rev_structure.items()}
rev_exceptions_structure = {i[0]: i[1].worksheets[0] for i in rev_exceptions_structure.items()}


# In[ ]:


# создадим словарь для каждого файла, где ключ-номер строки, а значение-уровень группировки
rev_structure = {i[0]: {row: i[1].row_dimensions[row].outline_level for row in i[1].row_dimensions} for i in rev_structure.items()}
rev_exceptions_structure = {i[0]: {row: i[1].row_dimensions[row].outline_level for row in i[1].row_dimensions} for i in rev_exceptions_structure.items()}


# In[ ]:


# преобразуем словари в pd.Series
rev_structure = {i[0]: pd.Series(i[1], name='Уровень').reset_index(drop=True) for i in rev_structure.items()}
rev_exceptions_structure = {i[0]: pd.Series(i[1], name='Уровень').reset_index(drop=True) for i in rev_exceptions_structure.items()}


# In[ ]:


# добавим в основные датафреймы уровни группировки строк
rev = {i[0]: i[1].merge(rev_structure[i[0]], left_index=True, right_index=True, how='inner') for i in rev.items()}
rev_exceptions = {i[0]: i[1].merge(rev_exceptions_structure[i[0]], left_index=True, right_index=True, how='inner') for i in rev_exceptions.items()}


# In[ ]:


rev = {i[0]: i[1].drop(list(range(8))) for i in rev.items()}
rev_exceptions = {i[0]: i[1].drop(list(range(6))) for i in rev_exceptions.items()}


# In[ ]:


for i in rev.values():
    i['Организация'] = i.apply(lambda row: row[0] if row['Уровень'] == 0 else np.nan,
                               axis=1)
    i['Счет'] = i.apply(lambda row: row[0] if row['Уровень'] == 1 else np.nan,
                        axis=1)
    i['Кор_счет'] = i.apply(lambda row: row[0] if row['Уровень'] == 2 else np.nan,
                            axis=1)
    i['Вид статьи'] = i.apply(lambda row: row[0] if row['Уровень'] == 3 else np.nan,
                              axis=1)
    i['Статья'] = i.apply(lambda row: row[0] if row['Уровень'] == 4 else np.nan,
                          axis=1)
    i['Дата'] = i.apply(lambda row: row[0] if row['Уровень'] == 5 else np.nan,
                        axis=1)
    i['Документ'] = i.apply(lambda row: row[0] if row['Уровень'] == 6 else np.nan,
                            axis=1)
    i['Операция'] = i.apply(lambda row: row[0] if row['Уровень'] == 7 else np.nan,
                            axis=1)


# In[ ]:


for i in rev_exceptions.values():
    i['Организация'] = i.apply(lambda row: row[0] if row['Уровень'] == 0 else np.nan,
                               axis=1)
    i['Счет'] = i.apply(lambda row: row[0] if row['Уровень'] == 1 else np.nan,
                        axis=1)
    i['Кор_счет'] = i.apply(lambda row: row[0] if row['Уровень'] == 2 else np.nan,
                            axis=1)
    i['Дата'] = i.apply(lambda row: row[0] if row['Уровень'] == 3 else np.nan,
                        axis=1)
    i['Документ'] = i.apply(lambda row: row[0] if row['Уровень'] == 4 else np.nan,
                            axis=1)
    i['Операция'] = i.apply(lambda row: row[0] if row['Уровень'] == 5 else np.nan,
                            axis=1)


# In[ ]:


rev = {i[0]: i[1].rename(columns={1: 'Изменение'}) for i in rev.items()}
rev_exceptions = {i[0]: i[1].rename(columns={1: 'Изменение'}) for i in rev_exceptions.items()}


# In[ ]:


# словари с именами столбцов и типами данных
type_of_columns_rev = {'Изменение': 'float64',
                       'Организация': 'object',
                       'Счет': 'object',
                       'Кор_счет': 'object', 
                       'Вид статьи': 'object',
                       'Статья': 'object',
                       'Документ': 'object',
                       'Операция': 'object'}
type_of_columns_rev_exc = {'Изменение': 'float64',
                           'Организация': 'object',
                           'Счет': 'object',
                           'Кор_счет': 'object',
                           'Документ': 'object',
                           'Операция': 'object'}


# In[ ]:


# подготовим строки для преобразования их в числовые значения
for i in rev.values():
    i['Изменение'] = i['Изменение'].apply(lambda row: str(row).replace(',', '.').replace(' ', '')
                                          if pd.notna(row)
                                          else row)
for i in rev_exceptions.values():
    i['Изменение'] = i['Изменение'].apply(lambda row: str(row).replace(',', '.').replace(' ', '')
                                          if pd.notna(row)
                                          else row)


# In[ ]:


# изменим тип данных в каждом датафрейме в соответствии со словарём 'type_of_columns'
rev = {i[0]: i[1].astype(type_of_columns_rev) for i in rev.items()}
rev_exceptions = {i[0]: i[1].astype(type_of_columns_rev_exc) for i in rev_exceptions.items()}


# In[ ]:


# заполним пропущенные значения в столбцах значениями предыдущей заполненной ячейки по строкам
for i in rev.values():
    i['Организация'] = i['Организация'].fillna(method='ffill', axis=0)
    i['Счет'] = i['Счет'].fillna(method='ffill', axis=0)
    i['Кор_счет'] = i['Кор_счет'].fillna(method='ffill', axis=0)
    i['Вид статьи'] = i['Вид статьи'].fillna(method='ffill', axis=0)
    i['Статья'] = i['Статья'].fillna(method='ffill', axis=0)
    i['Дата'] = i['Дата'].fillna(method='ffill', axis=0)
for i in rev_exceptions.values():
    i['Организация'] = i['Организация'].fillna(method='ffill', axis=0)
    i['Счет'] = i['Счет'].fillna(method='ffill', axis=0)
    i['Кор_счет'] = i['Кор_счет'].fillna(method='ffill', axis=0)
    i['Дата'] = i['Дата'].fillna(method='ffill', axis=0)


# In[ ]:


for i in rev.values():
    i.dropna(subset=['Операция'], inplace=True)
for i in rev_exceptions.values():
    i.dropna(subset=['Операция'], inplace=True)


# In[ ]:


date_formats = ['%d.%m.%Y', '%m/%d/%Y'] # форматы дат
# преобразуем строки в даты, используя один из форматов
for i in rev.values():
    i['Дата'] = i['Дата'].apply(lambda row: pd.to_datetime(row, format=next((f for f in date_formats if pd.to_datetime(row, format=f, errors='coerce') is not pd.NaT), None)))
for i in rev_exceptions.values():
    i['Дата'] = i['Дата'].apply(lambda row: pd.to_datetime(row, format=next((f for f in date_formats if pd.to_datetime(row, format=f, errors='coerce') is not pd.NaT), None)))


# In[ ]:


rev = {i[0]: i[1].drop(columns=[0, 'Уровень', 'Документ'])
       for i in rev.items()}
rev_exceptions = {i[0]: i[1].drop(columns=[0, 'Уровень', 'Документ'])
                  for i in rev_exceptions.items()}


# In[ ]:


# объединим счета 90.03, 90.04 и 90.05 со счётом 90.01, чтобы вычесть их из данного счёта
rev = {i[0]: pd.merge(i[1], rev_exceptions[i[0].replace('Выручка', 'НДС-акцизы-экспортные пошлины')],
                      on=['Организация', 'Дата', 'Операция'],
                      how='left', suffixes=('', '_exceptions')) for i in rev.items()}


# In[ ]:


for i in rev.values():
    i['Изменение'] = i['Изменение'] - i['Изменение_exceptions'].fillna(0)


# In[ ]:


rev = {i[0]: i[1].drop(columns=['Кор_счет', 'Операция', 'Изменение_exceptions', 'Счет_exceptions', 'Кор_счет_exceptions'])
       for i in rev.items()}


# In[ ]:


# исключим пустые датафреймы
rev = {i[0]: i[1] for i in rev.items() if not i[1].empty}


# In[ ]:


for i in rev.values():
    i['Статья'] = i['Статья'].fillna('-')


# In[ ]:


for i in rev.values():
    i['Статья'] = i['Статья'].apply(lambda row: f"{row.rsplit(',', 1)[0]} (ИНН: {row.rsplit(',', 1)[-1]})"
                                    if row.rsplit(',', 1)[-1] != ' '
                                    else row[:-2])


# In[ ]:


rev = {i[0]: i[1].groupby(['Организация',
                           'Вид статьи',
                           'Статья',
                           'Дата'], as_index=False)['Изменение'].sum() for i in rev.items()}


# In[ ]:


# создадим словарь, где ключ - название группы компаний, а значение - список таблиц по всем компаниям группы
rev_groups = {}

for key, value in rev.items():
    group = key.split('(')[1].split(')')[0]
    if group in rev_groups:
        rev_groups[group].append(value)
    else:
        rev_groups[group] = [value]


# In[ ]:


# создадим новый словарь с объединёнными таблицами
rev_merged = {f'Выручка_{i[0]}.xlsx': pd.concat(i[1], axis=0).reset_index(inplace=False, drop=True)
              for i in rev_groups.items()}


# In[ ]:


# определим последнюю дату по всем датафреймам группы компаний, чтобы продлить датафреймы с более ранними последними датами до самой поздней даты
last_date = {i[0]: i[1].sort_values(by='Дата')['Дата'].iloc[-1] for i in rev_merged.items()}


# In[ ]:


# создадим словарь со строками на последнюю дату из статей, которые нужно продлить
rev_new_rows = {i[0]: i[1].groupby(['Организация', 'Вид статьи', 'Статья'], as_index=False)['Дата'].max() for i in rev.items()}


# In[ ]:


# удалим статьи, которые не нужно продлять
for i in rev_new_rows.items():
    i[1].drop(i[1][i[1]['Дата'] == last_date[f'Выручка_{i[0].split("(")[1].split(")")[0]}.xlsx']].index, inplace=True)


# In[ ]:


# изменим добавляемые строки
for i in rev_new_rows.items():
    i[1]['Изменение'] = 0
    i[1]['Дата'] = last_date[f'Выручка_{i[0].split("(")[1].split(")")[0]}.xlsx']


# In[ ]:


# добавим новые строки в датафреймы
rev = {i[0]: pd.concat([i[1], rev_new_rows[i[0]]]) for i in rev.items()}


# In[ ]:


# создадим столбцы для добавления новых данных по остаткам на счетах
for i in rev.values():
    i['Разница'] = i.groupby(['Организация', 'Вид статьи', 'Статья'])['Дата'].diff()


# In[ ]:


# добавим промежуточные значения по остаткам на счёте, используя диапазоны дат
for i in rev.values():
    i['Начальная дата'] = i['Дата'] - i['Разница'] + timedelta(days=1)
    i['Конечная дата'] = i['Дата']
    i['Дата'] = i.apply(lambda row: pd.date_range(start=row['Начальная дата'],
                                                  end=row['Конечная дата']).tolist()
                        if pd.notna(row['Начальная дата'])
                        else row['Дата'],
                        axis=1)


# In[ ]:


# добавим новые строки с датами
rev = {i[0]: i[1].explode('Дата') for i in rev.items()}


# In[ ]:


for i in rev.values():
    i['Изменение'] = i.apply(lambda row: 0 if row['Дата'] != row['Конечная дата'] else row['Изменение'],
                             axis=1)


# In[ ]:


rev = {i[0]: i[1].drop(columns=['Разница',
                                'Начальная дата',
                                'Конечная дата'])
       for i in rev.items()}


# In[ ]:


# создадим словарь для добавления промежуточных значений по остаткам на счетах, используя данные по остаткам на дату накопительным итогом
# сгруппируем до уровня дат и получим уникальные значения статей для каждой даты
rev_values = {i[0]: i[1].groupby(['Организация', 'Вид статьи', 'Дата'], as_index=False)['Статья'].unique() for i in rev.items()} 


# In[ ]:


# отсортируем датафреймы по дате
rev_values = {i[0]: i[1].groupby(['Организация', 'Вид статьи'], as_index=False).apply(lambda row: row.sort_values('Дата')) for i in rev_values.items()} 


# In[ ]:


# присвоим индексы
rev_values = {i[0]: i[1].groupby(['Организация', 'Вид статьи'], as_index=False).apply(lambda row: row.reset_index(drop=True)) for i in rev_values.items()}


# In[ ]:


# добавим на каждую дату список из списков уникальных статей накопительным итогом
grouped_dict = {}
for key, value in rev_values.items():
    group_list = []
    for name, group in value.groupby(level=0):
        group['Статья (доп)'] = group.index.map(lambda row: group.loc[:row, 'Статья'].tolist())
        group['Статья (доп)'] = group['Статья (доп)'].apply(lambda row: set([x for sublist in row for x in sublist]))
        group_list.append(group)
    table = pd.concat(group_list, axis=0)
    grouped_dict[key] = table


# In[ ]:


rev_values = grouped_dict


# In[ ]:


# добавим новые строки со статьями
rev_values = {i[0]: i[1].explode('Статья (доп)') for i in rev_values.items()} 


# In[ ]:


for i in rev_values.values():
    i['Статья'] = i['Статья (доп)']
    i['Изменение'] = 0.0


# In[ ]:


rev_values = {i[0]: i[1].drop(columns=['Статья (доп)'])
              for i in rev_values.items()}


# In[ ]:


rev = {i[0]: pd.concat([i[1], rev_values[i[0]]]) for i in rev.items()}


# In[ ]:


rev = {i[0]: i[1].sort_values(by='Дата').groupby(['Организация',
                                                  'Вид статьи',
                                                  'Статья',
                                                  'Дата'],
                                                 as_index=False)['Изменение'].sum() for i in rev.items()}


# In[ ]:


for i in rev.values():
    i['Начальный остаток'] = 0.0
    i['Конечный остаток'] = 0.0


# In[ ]:


# создадим новые таблицы для расчёта остатков
rev_grouped = {i[0]: i[1].sort_values(by='Дата').groupby(['Организация',
                                                          'Вид статьи',
                                                          'Статья',
                                                          'Дата'],
                                                         as_index=False)['Изменение'].sum() for i in rev.items()}


# In[ ]:


grouped_dict = {}
for key, value in rev_grouped.items():
    # применение операций shift и cumsum к каждому счёту, организации и виду статьи отдельно
    group_list = []
    for name, group in value.groupby(['Организация',
                                      'Вид статьи',
                                      'Статья']):
        group['Начальный остаток'] = group['Изменение'].shift(fill_value=0).cumsum()
        group['Конечный остаток'] = group['Начальный остаток'] + group['Изменение']
        group_list.append(group)
    table = pd.concat(group_list, axis=0)
    grouped_dict[key] = table


# In[ ]:


rev_grouped = grouped_dict


# In[ ]:


merged_dict = {}
# объединим словарь с таблицами для расчёта остатков
for key, table1 in rev.items():
    table2 = rev_grouped.get(key)
    merged_table = pd.merge(table1, table2, on=['Организация',
                                                'Вид статьи',
                                                'Статья',
                                                'Дата'],
                            how='left',
                            suffixes=('', '_new'))
    merged_dict[key] = merged_table


# In[ ]:


rev = merged_dict


# In[ ]:


for i in rev.values():
    i['Начальный остаток'] = i['Начальный остаток_new']
    i['Конечный остаток'] = i['Конечный остаток_new']


# In[ ]:


# удалим лишние столбцы
rev = {i[0]: i[1].iloc[:, :-3] for i in rev.items()}


# In[ ]:


for i in rev.values():
    i['Показатель'] = 'Выручка'


# In[ ]:


rev = {i[0]: i[1].reset_index(drop=True) for i in rev.items()}


# In[ ]:


# создадим словарь, где ключ - название группы компаний, а значение - список таблиц по всем компаниям группы
rev_groups = {}

for key, value in rev.items():
    group = key.split('(')[1].split(')')[0]
    if group in rev_groups:
        rev_groups[group].append(value)
    else:
        rev_groups[group] = [value]


# In[ ]:


# создадим новый словарь с объединёнными таблицами
rev_merged = {f'Выручка_{i[0]}.xlsx': pd.concat(i[1], axis=0).reset_index(inplace=False, drop=True)
              for i in rev_groups.items()}


# In[ ]:


for i in rev_merged.items():
    print(f'{i[0]}:')
    print(i[1].info(show_counts=True))
    print()


# In[ ]:


# удалим исходные данные с расширением .xlsx из первоначальной папки
for i in source_rev_name:
    os.remove(os.path.join(source, i))
for i in source_rev_exceptions_name:
    os.remove(os.path.join(source, i))

